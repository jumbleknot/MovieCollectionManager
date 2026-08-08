"""T022: integration test — parse_spreadsheet against a REAL transient store (Redis).

Exercises the actual store seam the MCP tool uses (`store.read_upload`) end-to-end with
parser.parse_workbook — no store mocking (constitution §Test Type Integrity). Mirrors the
production flow: the BFF stashes upload bytes under `import:file:<handle>`; the parse tool
fetches them once (single-use) and structurally extracts tabs. Skips if Redis is absent.
"""

from __future__ import annotations

import os
import uuid
from pathlib import Path

import pytest
import redis.asyncio as redis
from openpyxl import load_workbook

from src import parser, store

REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379")
SAMPLE_XLSX = Path(__file__).resolve().parents[4] / "docs" / "test-data" / "sample-movies.xlsx"


def _expected_data_rows(sheet_name: str) -> int:
    """Count the fixture's data rows independently of the parser (048 FR-009 / US3-AC1).

    This assertion used to be a hard-coded `== 200` while the workbook has held **204** data rows
    since the fixture, the parser and the test all landed in one commit (33eba4c, 2026-06-14) — the
    test had never once run, so nothing regressed and nothing caught it. A literal is exactly the
    thing that rots silently when someone appends a row, so the expectation is derived from the
    fixture instead and moves with it.

    "Data row" is the spec's definition, not the parser's implementation: row 1 is the header, and a
    row counts if any cell in it is non-blank. That is deliberately re-derived here rather than
    imported from `parser`, which would make the assertion tautological — a parser that miscounted
    would then agree with itself.
    """
    workbook = load_workbook(SAMPLE_XLSX, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        return sum(
            1
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if any(str(cell).strip() for cell in row if cell is not None)
        )
    finally:
        workbook.close()


async def _redis_or_skip() -> redis.Redis:
    client = redis.from_url(REDIS_URL, decode_responses=False)
    try:
        await client.ping()
    except Exception:  # noqa: BLE001 — any connection failure → skip, this is an env gate
        await client.aclose()
        pytest.skip("Redis not available for spreadsheet-mcp integration test")
    return client


async def test_parse_via_real_transient_store_is_single_use() -> None:
    client = await _redis_or_skip()
    handle = uuid.uuid4().hex
    try:
        await client.set(store.IMPORT_PREFIX + handle, SAMPLE_XLSX.read_bytes(), ex=60)

        data = await store.read_upload(handle)
        result = parser.parse_workbook(data, "sample-movies.xlsx")

        sample = next(t for t in result["tabs"] if t["name"] == "Sample")
        assert sample["eligible"] is True
        expected_rows = _expected_data_rows("Sample")
        assert expected_rows > 0, "fixture workbook has no data rows — wrong file?"
        assert sample["rowCount"] == expected_rows

        # Single-use: a second read of the same handle fails (key deleted after first read).
        with pytest.raises(store.HandleNotFoundError):
            await store.read_upload(handle)
    finally:
        await client.delete(store.IMPORT_PREFIX + handle)
        await client.aclose()


async def test_missing_handle_raises_not_found() -> None:
    client = await _redis_or_skip()
    try:
        with pytest.raises(store.HandleNotFoundError):
            await store.read_upload("does-not-exist-" + uuid.uuid4().hex)
    finally:
        await client.aclose()
