"""T041: unit matrix for builder.build_workbook_bytes (US3 spreadsheet-mcp export).

Drives the pure workbook builder directly (no MCP transport, no Redis store). One sheet per
tab, header row from the ordered columns, multi-value cells `|`-joined, header-only sheet for a
collection with no movies, Excel-safe + de-duplicated sheet names, empty `tabs` rejected.

Covers: US3-AC2/3/4, FR-025/026/027.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from src.builder import WorkbookBuildError, build_workbook_bytes


def _reopen(data: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def _rows(ws: object) -> list[tuple[object, ...]]:
    return [tuple(r) for r in ws.iter_rows(values_only=True)]  # type: ignore[attr-defined]


# ---------------------------------------------------------------------------
# Structure: one sheet per tab, header row, returns openable bytes
# ---------------------------------------------------------------------------


def test_returns_openable_xlsx_bytes_and_filename() -> None:
    data, filename = build_workbook_bytes(
        [{"collectionName": "Favourites", "columns": ["Title", "Year"], "rows": []}]
    )
    assert isinstance(data, bytes) and data[:4] == b"PK\x03\x04"
    assert filename.endswith(".xlsx")
    _reopen(data)  # parses without error


def test_one_sheet_per_tab_named_by_collection() -> None:
    data, _ = build_workbook_bytes(
        [
            {"collectionName": "Action", "columns": ["Title"], "rows": []},
            {"collectionName": "Drama", "columns": ["Title"], "rows": []},
        ]
    )
    assert _reopen(data).sheetnames == ["Action", "Drama"]


def test_header_row_is_the_ordered_columns() -> None:
    data, _ = build_workbook_bytes(
        [{"collectionName": "C", "columns": ["Title", "Year", "Genres"], "rows": []}]
    )
    ws = _reopen(data)["C"]
    assert _rows(ws) == [("Title", "Year", "Genres")]


def test_rows_written_in_column_order_missing_attr_blank() -> None:
    data, _ = build_workbook_bytes(
        [
            {
                "collectionName": "C",
                "columns": ["Title", "Year", "Language"],
                "rows": [
                    {"Title": "Dune", "Year": "2021", "Language": "English"},
                    {"Title": "9", "Year": "2009"},  # no Language → blank cell
                ],
            }
        ]
    )
    assert _rows(_reopen(data)["C"]) == [
        ("Title", "Year", "Language"),
        ("Dune", "2021", "English"),
        ("9", "2009", None),
    ]


# ---------------------------------------------------------------------------
# Multi-value join (FR-027)
# ---------------------------------------------------------------------------


def test_list_cell_joined_with_default_pipe() -> None:
    data, _ = build_workbook_bytes(
        [
            {
                "collectionName": "C",
                "columns": ["Title", "Genres"],
                "rows": [{"Title": "X", "Genres": ["Action", "Sci-Fi", "Thriller"]}],
            }
        ]
    )
    assert _rows(_reopen(data)["C"])[1] == ("X", "Action|Sci-Fi|Thriller")


def test_list_cell_respects_custom_delimiter() -> None:
    data, _ = build_workbook_bytes(
        [
            {
                "collectionName": "C",
                "columns": ["Genres"],
                "rows": [{"Genres": ["A", "B"]}],
            }
        ],
        multi_value_delimiter=", ",
    )
    assert _rows(_reopen(data)["C"])[1] == ("A, B",)


def test_pre_joined_string_passes_through() -> None:
    data, _ = build_workbook_bytes(
        [
            {
                "collectionName": "C",
                "columns": ["Genres"],
                "rows": [{"Genres": "Action|Sci-Fi"}],
            }
        ]
    )
    assert _rows(_reopen(data)["C"])[1] == ("Action|Sci-Fi",)


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("=SUM(A1:A9)", "'=SUM(A1:A9)"),
        ("+1+1", "'+1+1"),
        ("-7", "'-7"),
        ("@cmd", "'@cmd"),
        ("The Matrix", "The Matrix"),  # safe value untouched
        ("'71", "'71"),  # legit leading apostrophe (not a trigger) untouched
    ],
)
def test_formula_trigger_cells_are_escaped(raw: str, expected: str) -> None:
    data, _ = build_workbook_bytes(
        [{"collectionName": "C", "columns": ["Title"], "rows": [{"Title": raw}]}]
    )
    assert _rows(_reopen(data)["C"])[1] == (expected,)


def test_empty_collection_yields_header_only_sheet() -> None:
    data, _ = build_workbook_bytes(
        [{"collectionName": "Empty", "columns": ["Title", "Year"], "rows": []}]
    )
    assert _rows(_reopen(data)["Empty"]) == [("Title", "Year")]


def test_empty_tabs_rejected() -> None:
    with pytest.raises(WorkbookBuildError):
        build_workbook_bytes([])


def test_duplicate_collection_names_are_de_duplicated() -> None:
    data, _ = build_workbook_bytes(
        [
            {"collectionName": "Watchlist", "columns": ["Title"], "rows": []},
            {"collectionName": "Watchlist", "columns": ["Title"], "rows": []},
        ]
    )
    names = _reopen(data).sheetnames
    assert len(names) == 2
    assert names[0] == "Watchlist"
    assert names[1] != "Watchlist"  # de-duplicated suffix


def test_invalid_sheet_name_chars_sanitized() -> None:
    data, _ = build_workbook_bytes(
        [{"collectionName": "Sci-Fi: [2024]/Best?", "columns": ["Title"], "rows": []}]
    )
    name = _reopen(data).sheetnames[0]
    assert not (set(name) & set(r"[]:*?/\\"))
    assert name  # non-empty


def test_long_sheet_name_truncated_to_31() -> None:
    long_name = "A" * 50
    data, _ = build_workbook_bytes(
        [{"collectionName": long_name, "columns": ["Title"], "rows": []}]
    )
    name = _reopen(data).sheetnames[0]
    assert len(name) <= 31


def test_long_names_colliding_after_truncation_stay_unique() -> None:
    base = "Collection With A Very Long Descriptive Name"
    data, _ = build_workbook_bytes(
        [
            {"collectionName": base + " One", "columns": ["Title"], "rows": []},
            {"collectionName": base + " Two", "columns": ["Title"], "rows": []},
        ]
    )
    names = _reopen(data).sheetnames
    assert len(set(names)) == 2
    assert all(len(n) <= 31 for n in names)
